from PyQt5.QtWidgets import QWidget, QMainWindow, QPushButton, QGroupBox, QTabWidget, QLabel, QMessageBox
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from openpyxl import load_workbook
import pyrebase
import math

# Icon File Directory
image_path = r"C:\Users\ghaut\Documents\PyCharm\firebase_to_py\logo\agv.png"

# Database Configuration
config = {
    "apiKey": "MrpgpaDPApr47c6ocnEkU1EX4B35nSHccrbo9RqP",
    "authDomain": "agv-data-acquisition.firebaseapp.com",
    "databaseURL": "https://agv-data-acquisition-default-rtdb.asia-southeast1.firebasedatabase.app/",
    "projectId": "agv-data-acquisition",
    "storageBucket": "agv-data-acquisition.appshot.com",
    "messagingSenderId": "238818806378",
}


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.count = 1
        self.trackpoint = []

        self.db = pyrebase.initialize_app(config).database()
        self.wb = load_workbook('Book1.xlsx')

        # Window Setup
        self.setObjectName("MainWindow")
        self.resize(640, 320)
        self.setMinimumSize(QSize(1200, 600))
        self.setMaximumSize(QSize(1200, 600))
        self.setBaseSize(QSize(0, 0))
        self.setAutoFillBackground(True)
        self.setTabShape(QTabWidget.Triangular)
        self.setWindowTitle("AGV Monitoring System")

        # Central Widget Setup
        self.centralwidget = QWidget(self)
        self.centralwidget.setObjectName("centralwidget")
        self.setCentralWidget(self.centralwidget)

        # Border Widget Setup
        self.widget = QWidget(self.centralwidget)
        self.widget.setGeometry(QRect(0, 0, 270, 600))
        self.widget.setAutoFillBackground(True)
        self.widget.setObjectName("widget")

        self.widgetup = QWidget(self.centralwidget)
        self.widgetup.setGeometry(QRect(270, 0, 930, 20))
        self.widgetup.setAutoFillBackground(True)
        self.widgetup.setObjectName("widgetup")

        self.widgetright = QWidget(self.centralwidget)
        self.widgetright.setGeometry(QRect(1181, 0, 20, 600))
        self.widgetright.setAutoFillBackground(True)
        self.widgetright.setObjectName("widgetright")

        self.widgetdown = QWidget(self.centralwidget)
        self.widgetdown.setGeometry(QRect(270, 581, 930, 20))
        self.widgetdown.setAutoFillBackground(True)
        self.widgetdown.setObjectName("widget")

        # Group Box Setup
        self.groupBox = QGroupBox(self.widget)
        self.groupBox.setGeometry(QRect(10, 10, 231, 580))
        self.groupBox.setAutoFillBackground(True)
        self.groupBox.setObjectName("groupBox")
        font = QFont()
        font.setPointSize(10)
        self.groupBox.setFont(font)
        self.groupBox.setTitle("Parameter")
        self.groupBox.raise_()

        # Stop Push Button Setup
        self.pushButton_2 = QPushButton(self.groupBox)
        self.pushButton_2.setGeometry(QRect(10, 530, 211, 41))
        font = QFont()
        font.setPointSize(10)
        self.pushButton_2.setFont(font)
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_2.clicked.connect(self.stop_button)
        self.pushButton_2.setText("Stop")

        # Start Push Button Setup
        self.pushButton = QPushButton(self.groupBox)
        self.pushButton.setGeometry(QRect(10, 480, 211, 41))
        font = QFont()
        font.setPointSize(10)
        self.pushButton.setFont(font)
        self.pushButton.setObjectName("pushButton")
        self.pushButton.clicked.connect(self.start_button)
        self.pushButton.setText("Start")

        # Clear Push Button Setup
        self.pushButton_3 = QPushButton(self.groupBox)
        self.pushButton_3.setGeometry(QRect(10, 400, 211, 41))
        font = QFont()
        font.setPointSize(10)
        self.pushButton_3.setFont(font)
        self.pushButton_3.setObjectName("pushButton_3")
        self.pushButton_3.clicked.connect(self.clear_button)
        self.pushButton_3.setText("Clear Trace")

        # X label and X Value
        self.x_label = QLabel(self.groupBox)
        self.x_label.setGeometry(QRect(10, 30, 61, 31))
        self.x_label.setTextFormat(Qt.AutoText)
        self.x_label.setObjectName("x_label")
        self.x_label.setText(
            "<html><head/><body><p><span style=\" font-size:10pt;\">x   :</span></p></body></html>")

        self.x_val = QLabel(self.groupBox)
        self.x_val.setGeometry(QRect(130, 30, 91, 31))
        self.x_val.setObjectName("x_val")
        self.x_val.setText(
            "<html><head/><body><p align=\"right\"><span style=\" font-size:10pt;\">0</span></p></body></html>")

        # Y Label and Y Value
        self.y_label = QLabel(self.groupBox)
        self.y_label.setGeometry(QRect(10, 70, 61, 31))
        self.y_label.setObjectName("y_label")
        self.y_label.setText(
            "<html><head/><body><p><span style=\" font-size:10pt;\">y :</span></p></body></html>")

        self.y_val = QLabel(self.groupBox)
        self.y_val.setGeometry(QRect(130, 70, 91, 31))
        self.y_val.setObjectName("y_val")
        self.y_val.setText(
            "<html><head/><body><p align=\"right\"><span style=\" font-size:10pt;\">0</span></p></body></html>")

        # Angle Label and Value
        self.angle_label = QLabel(self.groupBox)
        self.angle_label.setGeometry(QRect(10, 110, 61, 31))
        self.angle_label.setObjectName("angle_label")
        self.angle_label.setText(
            "<html><head/><body><p><span style=\" font-size:10pt;\">angle :</span></p></body></html>")

        self.angle_val = QLabel(self.groupBox)
        self.angle_val.setGeometry(QRect(130, 110, 91, 31))
        self.angle_val.setAlignment(Qt.AlignRight | Qt.AlignTrailing | Qt.AlignVCenter)
        self.angle_val.setObjectName("angle_val")
        self.angle_val.setText(
            "<html><head/><body><p><span style=\" font-size:10pt;\">0</span></p></body></html>")

        # Variable Declaration
        self.icon_rotation = 0
        self.icon_pos = QPoint(0, 0)
        self.icon_size = QSize(50, 60)
        self.icon_attr = QRect(self.icon_pos, self.icon_size)

        # Timer Initialization
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_label)

        self.timer_2 = QTimer(self)
        self.timer_2.timeout.connect(self.upload_data)
        self.timer_2.start(500)

    def upload_data(self):
        ws1 = self.wb['circle']
        datax = ws1.cell(row=self.count, column=1)
        ws2 = self.wb['circle']
        datay = ws2.cell(row=self.count, column=2)
        ws3 = self.wb['circle']
        datang = ws3.cell(row=self.count, column=3)
        if datax.value is None and datay.value is None and datang.value is None:
            self.count = 1
        elif datax.value and datay.value:
            if datang.value is None:
                self.count = 1
            else:
                try:
                    self.db.child("test").update({"X": int(datax.value)})
                    self.db.child("test").update({"Y": int(datay.value)})
                    self.db.child("test").update({"angle": int(datang.value)})
                except Exception:
                    pass
                self.count = self.count + 1
        else:
            self.count = 1

    def update_label(self):
        # Retrieving data from firebase
        try:
            x = self.db.child("test").child("X").get().val()
            y = self.db.child("test").child("Y").get().val()
            ang = self.db.child("test").child("angle").get().val()
        except Exception:
            dbmsg = QMessageBox()
            dbmsg.setIcon(QMessageBox.Warning)
            dbmsg.setWindowTitle("Error")
            dbmsg.setText("Please check your internet connection or your server configuration!")
            dbmsg.setStandardButtons(QMessageBox.Ok)
            dbmsg.buttonClicked.connect(self.msg_button)
            dbmsg.exec_()
            return

        self.x_val.setText(
            f'<html><head/><body><p align=\"right\"><span style=\" font-size:10pt;\">{str(x)}</span></p></body></html>')
        self.y_val.setText(
            f'<html><head/><body><p align=\"right\"><span style=\" font-size:10pt;\">{str(y)}</span></p></body></html>')
        self.angle_val.setText(
            f"<html><head/><body><p><span style=\" font-size:10pt;\">{str(ang)}</span></p></body></html>")

        # Movement Update
        if (int(x)) <= 270:
            pos_x = 270 - (self.icon_attr.width() / 2)
            self.icon_pos.setX(int(pos_x))
        elif (int(x)) >= 1180:
            pos_x = 1180 - (self.icon_attr.width() / 2)
            self.icon_pos.setX(int(pos_x))
        else:
            pos_x = (int(x)) - (self.icon_attr.width() / 2)
            self.icon_pos.setX(int(pos_x))

        if (int(y)) <= 20:
            pos_y = 20 - (self.icon_attr.height() / 2)
            self.icon_pos.setY(int(pos_y))
        elif (int(y)) >= 560:
            pos_y = 560 - (self.icon_attr.height() / 2)
            self.icon_pos.setY(int(pos_y))
        else:
            pos_y = (int(y)) - (self.icon_attr.height() / 2)
            self.icon_pos.setY(int(pos_y))

        # Rotation Update
        self.icon_rotation = int(ang)

        # Scaling Update
        rad = (self.icon_rotation * math.pi) / 180
        h = int(abs((abs(math.sin(rad) * 50.0)) + (abs(math.cos(rad) * 60.0))))
        w = int(abs((abs(math.cos(rad) * 50.0)) + (abs(math.sin(rad) * 60.0))))
        self.icon_size.setWidth(w)
        self.icon_size.setHeight(h)

        # Track Point Update
        self.trackpoint.append(QPointF(int(x), int(y)))

        # Update the monitor
        self.update()

    def paintEvent(self, event):
        paint = QPainter(self)

        # Background
        paint.setPen(QPen(Qt.white, 0, Qt.SolidLine))
        paint.setBrush(QBrush(Qt.white, Qt.SolidPattern))
        rec = QRect(270, 20, 910, 560)
        paint.drawRect(rec)

        # Grid
        horline = QPainter(self)
        horline.setPen(QPen(Qt.lightGray, 1, Qt.SolidLine))
        x = 270
        for i in range(92):
            horline.drawLine(x, 20, x, 580)
            x = x + 10

        verline = QPainter(self)
        verline.setPen(QPen(Qt.lightGray, 1, Qt.SolidLine))
        y = 20
        for i in range(57):
            verline.drawLine(270, y, 1180, y)
            y = y + 10

        # Track
        track = QPainter(self)
        track.setPen(QPen(Qt.black, 4, Qt.SolidLine))
        track.drawLine(730, 430, 730, 130)
        track.drawLine(730, 130, 880, 130)
        track.drawArc(800, 130, 150, 150, 0*16, 90*16)
        track.drawLine(950, 200, 950, 280)
        track.drawLine(950, 280, 510, 280)
        track.drawLine(510, 280, 510, 360)
        track.drawArc(510, 280, 150, 150, -90 * 16, -90 * 16)
        track.drawLine(580, 430, 730, 430)

        # AGV Icon
        icon = QPixmap(image_path)
        self.icon_attr = QRect(self.icon_pos, self.icon_size)
        transform = QTransform()
        transform.rotate(self.icon_rotation)
        icon = icon.transformed(transform, mode=Qt.SmoothTransformation)
        paint.drawPixmap(self.icon_attr, icon)

        # Trace Point
        point = QPainter(self)
        point.setPen(QPen(Qt.red, 8))
        for coord in self.trackpoint:
            point.drawPoint(coord)

    def start_button(self):
        self.timer.start(500)
        print("Starting...")

    def stop_button(self):
        self.timer.stop()
        self.x_val.setText(
            '<html><head/><body><p align=\"right\"><span style=\" font-size:10pt;\">0</span></p></body></html>')
        self.y_val.setText(
            '<html><head/><body><p align=\"right\"><span style=\" font-size:10pt;\">0</span></p></body></html>')
        self.angle_val.setText(
            "<html><head/><body><p><span style=\" font-size:10pt;\">0</span></p></body></html>")

    def msg_button(self):
        self.timer.stop()

    def clear_button(self):
        self.trackpoint = []
        self.update()
